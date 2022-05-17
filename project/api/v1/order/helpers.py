import os
import logging
import openpyxl

from django.conf import settings
from rest_framework.request import Request

from typing import Tuple
from openpyxl.styles import NamedStyle, Side, Border, Font

from apps.order.models import Order
from apps.cart.models import Cart

logger = logging.getLogger("django")


def get_order_info_for_email(order: Order, cart: Cart) -> Tuple[str, str]:
    """ Функиця формирует инфо о заказе для отправке по
    электронной почте, возвращает кортеж для html и текста.
    """
    text = f"""
Тип доставки: {'Самовывоз' if order.delivery_type == Order.DeliveryType.pickup else 'Транспортная компания'}\n
Инициалы: {order.fio}\n
Почта: {order.email}\n
Телефон: {order.phone}\n
ИНН: {order.inn}\n
Компания: {order.name_company}\n\n\n
    """
    html = text = f"""
Тип доставки: {'Самовывоз' if order.delivery_type == Order.DeliveryType.pickup else 'Транспортная компания'}<br>
Инициалы: {order.fio}<br>
Почта: {order.email}<br>
Телефон: {order.phone}<br>
ИНН: {order.inn}<br>
Компания: {order.name_company}<br><br><br>
    """

    for idx, relation in enumerate(cart.cart_products.all(), start=1):
        vendor_code = relation.product.vendor_code
        fullname_with_attrs = relation.product.fullname_with_attrs
        text += f"{idx}. {vendor_code} {fullname_with_attrs}; {relation.count} шт.\n"
        html += f"{idx}. {vendor_code} {fullname_with_attrs}; {relation.count} шт.<br>"
    return text, html


def check_excel_orders_dir() -> None:
    """ Функция проверяет наличие директории с файлами заказов.
    """
    if not os.path.exists(f"{settings.MEDIA_ROOT}/order_excels"):
        os.mkdir(f"{settings.MEDIA_ROOT}/order_excels")


def get_implementer_name_by_request(request: Request) -> str:
    """ Фукнция возвращает значение для поля "Исполнитель" в зависимости
    от филиала, с которого делается запрос.
    """
    try:
        http_origin = request.META['HTTP_ORIGIN']
    except KeyError:
        http_origin = 'afinara.ru'

    if "a-mitra" in http_origin:
        return "ООО «Амитра»"
    elif "atlant-a" in http_origin:
        return "ООО «Атлант-А»"
    elif "a-sibra" in http_origin:
        return "ООО «А-Сибра»"
    elif "arlan" in http_origin:
        return "ООО «Арлан»"
    else:
        return "ООО ГК «Афинара-ПТ»"


def make_excel_about_order(order: Order, cart: Cart, request: Request) -> str:
    """ Функция формирует excel-файл с информацией о заказе,
    возвращает путь до него.
    """
    check_excel_orders_dir()

    wb = openpyxl.Workbook()
    wb.create_sheet(title="Заказ", index=0)
    sheet = wb["Заказ"]

    border = Side(style="thin", color="000000")
    # Стиль для заголовков
    tstyle = NamedStyle(name="highlight")
    tstyle.font = Font(bold=True, size=12)
    tstyle.border = Border(left=border, top=border, right=border, bottom=border)
    wb.add_named_style(tstyle)

    # Стиль для списка товаров
    lstyle = NamedStyle(name="goodsstyle")
    lstyle.font = Font(size=11)
    lstyle.border = Border(left=border, top=border, right=border, bottom=border)
    wb.add_named_style(lstyle)

    # Объединить колонки для заголовка
    sheet.merge_cells("A1:H1")
    sheet["A1"] = f"Заказ поставщику от {order.created_at.strftime('%Y-%m-%d %H:%M')}"
    sheet["A1"].style = "highlight"
    sheet["B1"].style = "highlight"
    sheet["C1"].style = "highlight"
    sheet["D1"].style = "highlight"
    sheet["E1"].style = "highlight"
    sheet["F1"].style = "highlight"
    sheet["G1"].style = "highlight"
    sheet["H1"].style = "highlight"
    sheet.row_dimensions[1].height = 25

    # Объединить колонки для инфы об исполнителе
    sheet.merge_cells("A3:H3")
    sheet["A3"] = f"Исполнитель: {get_implementer_name_by_request(request)}"
    sheet["A3"].style = "highlight"
    sheet["B3"].style = "highlight"
    sheet["C3"].style = "highlight"
    sheet["D3"].style = "highlight"
    sheet["E3"].style = "highlight"
    sheet["F3"].style = "highlight"
    sheet["G3"].style = "highlight"
    sheet["H3"].style = "highlight"
    sheet.row_dimensions[3].height = 25

    # Объединить колонки для инфы о заказчике
    sheet.merge_cells("A5:H5")
    sheet["A5"] = f"Заказчик: {order.fio}, {order.phone}, {order.email}   ИНН: {order.inn}   {order.name_company}"
    sheet["A5"].style = "highlight"
    sheet["B5"].style = "highlight"
    sheet["C5"].style = "highlight"
    sheet["D5"].style = "highlight"
    sheet["E5"].style = "highlight"
    sheet["F5"].style = "highlight"
    sheet["G5"].style = "highlight"
    sheet["H5"].style = "highlight"
    sheet.row_dimensions[5].height = 25

    # Заголовки для таблицы товаров
    sheet["A7"] = "№"
    sheet["A7"].style = "highlight"
    sheet.column_dimensions["A"].width = 5

    sheet["B7"] = "Артикул"
    sheet["B7"].style = "highlight"
    sheet.column_dimensions["B"].width = 20

    sheet["C7"] = "Товары (работы, услуги)"
    sheet["C7"].style = "highlight"
    sheet.column_dimensions["C"].width = 70

    sheet["D7"] = "Вид цены"
    sheet["D7"].style = "highlight"
    sheet.column_dimensions["D"].width = 15

    sheet.merge_cells("E7:F7")
    sheet["E7"] = "Количество"
    sheet["E7"].style = "highlight"
    sheet["F7"].style = "highlight"
    sheet.column_dimensions["E"].width = 15

    sheet["G7"] = "Цена"
    sheet["G7"].style = "highlight"
    sheet.column_dimensions["F"].width = 15

    sheet["H7"] = "Сумма"
    sheet["H7"].style = "highlight"
    sheet.column_dimensions["G"].width = 15

    # Заполняем таблицу в цикле товарами, начало с 8 строки
    row_cntr = 8
    total_amount = 0
    for idx, relation in enumerate(cart.cart_products.all(), start=1):
        total_cost = relation.total_cost()
        params = [
            idx, relation.product.vendor_code, relation.product.fullname_with_attrs, "Розничная",
            relation.count, relation.product.measure, relation.product.price, total_cost,
        ]

        for col in range(1, 9):
            sheet.cell(row=row_cntr, column=col).value = params[col-1]
            sheet.cell(row=row_cntr, column=col).style = "goodsstyle"

        if total_cost:
            total_amount += total_cost

        row_cntr += 1

    sheet["G17"] = "Итого"
    sheet["G17"].style = "goodsstyle"
    sheet["H17"] = total_amount
    sheet["H17"].style = "goodsstyle"

    sheet["B19"] = "Менеджер"
    sheet["B19"].style = "goodsstyle"
    sheet["C19"].style = "goodsstyle"

    filedest = f"{settings.MEDIA_ROOT}/order_excels/order_{order.fio}_{order.created_at.strftime('%Y-%m-%d')}.xlsx"
    wb.save(filedest)
    return filedest
