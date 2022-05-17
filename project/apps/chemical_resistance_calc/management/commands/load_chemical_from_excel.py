import os
import logging

from django.core.management.base import BaseCommand

from openpyxl import load_workbook

from apps.chemical_resistance_calc.models import TableOfChemicalResistance

logger = logging.getLogger('django_info')


class Command(BaseCommand):
    help = 'Загрузить информацию из таблицы химических сопротивлений в БД.'

    def handle(self, *args, **kwargs):
        TableOfChemicalResistance.objects.all().delete()
        os.chdir('apps/chemical_resistance_calc/management/commands')
        wb = load_workbook('chemical_resistance_table.xlsx')
        ws = wb.active

        row_cntr = 29
        while True:
            name = ws.cell(row=row_cntr, column=1).value
            if not name:
                logger.info('Заполнение успешно завершено')
                break

            pvh = ws.cell(row=row_cntr, column=6).value
            if not pvh or pvh == '':
                pvh = None

            pp = ws.cell(row=row_cntr, column=7).value
            if not pp or pp == '':
                pp = None

            pvdf = ws.cell(row=row_cntr, column=9).value
            if not pvdf or pvdf == '':
                pvdf = None

            hpvh = ws.cell(row=row_cntr, column=10).value
            if not hpvh or hpvh == '':
                hpvh = None

            nbr = ws.cell(row=row_cntr, column=11).value
            if not nbr or nbr == '':
                nbr = None

            epdm = ws.cell(row=row_cntr, column=12).value
            if not epdm or epdm == '':
                epdm = None

            fmk = ws.cell(row=row_cntr, column=13).value
            if not fmk or fmk == '':
                fmk = None

            ptfe = ws.cell(row=row_cntr, column=14).value
            if not ptfe or ptfe == '':
                ptfe = None

            TableOfChemicalResistance.objects.create(**{
                'name': name,
                'state': ws.cell(row=row_cntr, column=2).value,
                'formula': ws.cell(row=row_cntr, column=3).value,
                'concentration': ws.cell(row=row_cntr, column=4).value,
                'temperature': ws.cell(row=row_cntr, column=5).value,
                'pvh': pvh,
                'pp': pp,
                'pvdf': pvdf,
                'hpvh': hpvh,
                'nbr': nbr,
                'epdm': epdm,
                'fmk': fmk,
                'ptfe': ptfe,
            })
            row_cntr += 1
